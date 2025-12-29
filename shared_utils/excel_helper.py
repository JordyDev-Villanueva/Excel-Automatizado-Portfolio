"""
Excel Helper - Utilidades Compartidas para Automatización de Excel
===================================================================
Módulo con funciones reutilizables para procesamiento, análisis y
formateo profesional de archivos Excel.

Autor: Excel Automatizado
Fecha: Diciembre 2025
"""

import logging
import sys
from pathlib import Path
from typing import Optional, Union, List, Tuple
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE COLORES CORPORATIVOS
# ═══════════════════════════════════════════════════════════════

COLORES = {
    'azul_oscuro': '1F4788',      # Títulos, headers
    'azul_claro': 'D6E4F5',       # Fondos KPIs
    'gris': '7F7F7F',             # Texto secundario
    'verde': '70AD47',            # Valores positivos
    'naranja': 'ED7D31',          # Alertas
    'blanco': 'FFFFFF',           # Texto en headers
}

# ═══════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════

def setup_logger(name: str, nivel: int = logging.INFO) -> logging.Logger:
    """
    Configura un logger profesional para el proyecto.

    Args:
        name (str): Nombre del logger (usualmente __name__)
        nivel (int): Nivel de logging (default: INFO)

    Returns:
        logging.Logger: Logger configurado

    Ejemplo:
        >>> logger = setup_logger(__name__)
        >>> logger.info("Proceso iniciado")
    """
    logger = logging.getLogger(name)
    logger.setLevel(nivel)

    # Evitar duplicados si ya existe un handler
    if not logger.handlers:
        # Handler para consola
        handler = logging.StreamHandler(sys.stdout)
        handler.setLevel(nivel)

        # Formato profesional
        formato = logging.Formatter(
            '%(asctime)s | %(levelname)-8s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        handler.setFormatter(formato)
        logger.addHandler(handler)

    return logger


# ═══════════════════════════════════════════════════════════════
# LECTURA Y PROCESAMIENTO DE ARCHIVOS
# ═══════════════════════════════════════════════════════════════

def leer_archivos_excel(
    carpeta_path: Union[str, Path],
    patron: str = "*.xlsx",
    columnas_requeridas: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    Lee todos los archivos Excel de una carpeta y los combina en un DataFrame.

    Args:
        carpeta_path: Ruta de la carpeta con archivos Excel
        patron: Patrón de archivos a buscar (default: "*.xlsx")
        columnas_requeridas: Lista de columnas que deben existir (opcional)

    Returns:
        pd.DataFrame: DataFrame consolidado con todos los datos

    Raises:
        FileNotFoundError: Si la carpeta no existe
        ValueError: Si no se encuentran archivos o faltan columnas

    Ejemplo:
        >>> df = leer_archivos_excel("input/", columnas_requeridas=['Fecha', 'Producto'])
        >>> print(f"Total registros: {len(df)}")
    """
    carpeta = Path(carpeta_path)
    logger = logging.getLogger(__name__)

    # Validar que la carpeta existe
    if not carpeta.exists():
        raise FileNotFoundError(f"La carpeta no existe: {carpeta}")

    # Buscar archivos Excel
    archivos = list(carpeta.glob(patron))

    if not archivos:
        raise ValueError(f"No se encontraron archivos {patron} en {carpeta}")

    logger.info(f"Encontrados {len(archivos)} archivo(s) para procesar")

    # Lista para almacenar DataFrames
    dataframes = []

    # Leer cada archivo
    for archivo in archivos:
        try:
            logger.info(f"  → Leyendo: {archivo.name}")
            df = pd.read_excel(archivo)

            # Validar columnas requeridas
            if columnas_requeridas:
                columnas_faltantes = set(columnas_requeridas) - set(df.columns)
                if columnas_faltantes:
                    raise ValueError(
                        f"El archivo {archivo.name} no tiene las columnas: {columnas_faltantes}"
                    )

            dataframes.append(df)
            logger.info(f"    ✓ {len(df)} registros cargados")

        except Exception as e:
            logger.error(f"    ✗ Error en {archivo.name}: {str(e)}")
            raise

    # Combinar todos los DataFrames
    df_consolidado = pd.concat(dataframes, ignore_index=True)

    logger.info(f"✓ Consolidación completa: {len(df_consolidado)} registros totales")

    return df_consolidado


# ═══════════════════════════════════════════════════════════════
# GENERACIÓN DE GRÁFICOS PROFESIONALES
# ═══════════════════════════════════════════════════════════════

def crear_grafico_barras(
    data: Union[pd.Series, pd.DataFrame],
    titulo: str,
    output_path: Union[str, Path],
    columna_x: Optional[str] = None,
    columna_y: Optional[str] = None,
    color: str = '#1F4788',
    figsize: Tuple[int, int] = (10, 6),
    formato_y: str = 'miles'
) -> None:
    """
    Genera un gráfico de barras profesional y lo guarda como PNG.

    Args:
        data: Series o DataFrame con los datos
        titulo: Título del gráfico
        output_path: Ruta donde guardar el gráfico
        columna_x: Nombre de columna para eje X (si data es DataFrame)
        columna_y: Nombre de columna para eje Y (si data es DataFrame)
        color: Color de las barras (default: azul oscuro)
        figsize: Tamaño de la figura en pulgadas (default: 10x6)
        formato_y: Formato del eje Y ('miles', 'millones', 'porcentaje')

    Ejemplo:
        >>> ventas = df.groupby('Sucursal')['Total'].sum()
        >>> crear_grafico_barras(ventas, "Ventas por Sucursal", "output/ventas.png")
    """
    plt.figure(figsize=figsize)
    sns.set_style("whitegrid")

    # Si es Series, usar directamente
    if isinstance(data, pd.Series):
        ax = data.plot(kind='bar', color=color, edgecolor='black', linewidth=0.5)
        plt.xlabel(data.index.name or '', fontsize=11, fontweight='bold')
        plt.ylabel('Valor', fontsize=11, fontweight='bold')

    # Si es DataFrame, usar columnas especificadas
    elif isinstance(data, pd.DataFrame):
        if not columna_x or not columna_y:
            raise ValueError("Para DataFrame, especifica columna_x y columna_y")
        ax = data.plot(x=columna_x, y=columna_y, kind='bar', color=color,
                       edgecolor='black', linewidth=0.5, legend=False)
        plt.xlabel(columna_x, fontsize=11, fontweight='bold')
        plt.ylabel(columna_y, fontsize=11, fontweight='bold')

    plt.title(titulo, fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')

    # Formato del eje Y
    if formato_y == 'miles':
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
    elif formato_y == 'millones':
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x/1e6:.1f}M'))
    elif formato_y == 'porcentaje':
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:.1f}%'))

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    logging.getLogger(__name__).info(f"  ✓ Gráfico guardado: {Path(output_path).name}")


def crear_grafico_circular(
    data: pd.Series,
    titulo: str,
    output_path: Union[str, Path],
    colores: Optional[List[str]] = None,
    figsize: Tuple[int, int] = (10, 6),
    mostrar_porcentaje: bool = True
) -> None:
    """
    Genera un gráfico circular (pie chart) profesional y lo guarda como PNG.

    Args:
        data: Series con los datos (índice = etiquetas, valores = cantidades)
        titulo: Título del gráfico
        output_path: Ruta donde guardar el gráfico
        colores: Lista de colores hexadecimales (opcional)
        figsize: Tamaño de la figura en pulgadas (default: 10x6)
        mostrar_porcentaje: Si mostrar porcentajes en las etiquetas

    Ejemplo:
        >>> categorias = df.groupby('Categoría')['Total'].sum()
        >>> crear_grafico_circular(categorias, "Distribución", "output/pie.png")
    """
    plt.figure(figsize=figsize)

    # Colores por defecto si no se especifican
    if colores is None:
        colores = ['#1F4788', '#70AD47', '#ED7D31', '#4472C4', '#FFC000', '#5B9BD5']

    # Formato de etiquetas
    def autopct_format(pct):
        return f'{pct:.1f}%' if pct > 3 else ''  # Solo mostrar si > 3%

    autopct = autopct_format if mostrar_porcentaje else None

    plt.pie(data, labels=data.index, autopct=autopct, startangle=90,
            colors=colores, textprops={'fontsize': 10})

    plt.title(titulo, fontsize=14, fontweight='bold', pad=20)
    plt.axis('equal')  # Círculo perfecto

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    logging.getLogger(__name__).info(f"  ✓ Gráfico guardado: {Path(output_path).name}")


def crear_grafico_linea(
    data: Union[pd.Series, pd.DataFrame],
    titulo: str,
    output_path: Union[str, Path],
    columna_x: Optional[str] = None,
    columna_y: Optional[str] = None,
    color: str = '#1F4788',
    figsize: Tuple[int, int] = (10, 6),
    marcar_puntos: bool = True
) -> None:
    """
    Genera un gráfico de línea para tendencias y lo guarda como PNG.

    Args:
        data: Series o DataFrame con los datos
        titulo: Título del gráfico
        output_path: Ruta donde guardar el gráfico
        columna_x: Nombre de columna para eje X (si data es DataFrame)
        columna_y: Nombre de columna para eje Y (si data es DataFrame)
        color: Color de la línea (default: azul oscuro)
        figsize: Tamaño de la figura en pulgadas (default: 10x6)
        marcar_puntos: Si mostrar marcadores en los puntos

    Ejemplo:
        >>> ventas_diarias = df.groupby('Fecha')['Total'].sum()
        >>> crear_grafico_linea(ventas_diarias, "Tendencia", "output/trend.png")
    """
    plt.figure(figsize=figsize)
    sns.set_style("whitegrid")

    marker = 'o' if marcar_puntos else None

    # Si es Series
    if isinstance(data, pd.Series):
        plt.plot(data.index, data.values, color=color, linewidth=2,
                marker=marker, markersize=6, markerfacecolor='white',
                markeredgewidth=2, markeredgecolor=color)
        plt.xlabel(data.index.name or 'Fecha', fontsize=11, fontweight='bold')
        plt.ylabel('Valor', fontsize=11, fontweight='bold')

    # Si es DataFrame
    elif isinstance(data, pd.DataFrame):
        if not columna_x or not columna_y:
            raise ValueError("Para DataFrame, especifica columna_x y columna_y")
        plt.plot(data[columna_x], data[columna_y], color=color, linewidth=2,
                marker=marker, markersize=6, markerfacecolor='white',
                markeredgewidth=2, markeredgecolor=color)
        plt.xlabel(columna_x, fontsize=11, fontweight='bold')
        plt.ylabel(columna_y, fontsize=11, fontweight='bold')

    plt.title(titulo, fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    logging.getLogger(__name__).info(f"  ✓ Gráfico guardado: {Path(output_path).name}")


# ═══════════════════════════════════════════════════════════════
# FORMATEO PROFESIONAL DE EXCEL
# ═══════════════════════════════════════════════════════════════

def aplicar_formato_profesional(
    worksheet,
    tiene_header: bool = True,
    autoajustar_columnas: bool = True,
    aplicar_bordes: bool = True
) -> None:
    """
    Aplica formato profesional estándar a una hoja de Excel.

    Args:
        worksheet: Objeto worksheet de openpyxl
        tiene_header: Si la primera fila es header (default: True)
        autoajustar_columnas: Si ajustar ancho de columnas automáticamente
        aplicar_bordes: Si aplicar bordes a las celdas

    Ejemplo:
        >>> wb = load_workbook('archivo.xlsx')
        >>> ws = wb['Hoja1']
        >>> aplicar_formato_profesional(ws)
        >>> wb.save('archivo.xlsx')
    """
    # Definir estilos
    header_font = Font(name='Calibri', size=11, bold=True, color=COLORES['blanco'])
    header_fill = PatternFill(start_color=COLORES['azul_oscuro'],
                              end_color=COLORES['azul_oscuro'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')

    border_side = Side(style='thin', color='D3D3D3')
    border = Border(left=border_side, right=border_side,
                   top=border_side, bottom=border_side)

    # Formatear header (fila 1)
    if tiene_header:
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            if aplicar_bordes:
                cell.border = border

        worksheet.row_dimensions[1].height = 25

    # Formatear celdas de datos
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    inicio_datos = 2 if tiene_header else 1

    for row in range(inicio_datos, max_row + 1):
        worksheet.row_dimensions[row].height = 20

        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = cell_alignment

            if aplicar_bordes:
                cell.border = border

    # Autoajustar ancho de columnas
    if autoajustar_columnas:
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)

            # Calcular ancho máximo
            max_length = 0
            for row in range(1, max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # Establecer ancho (con padding)
            adjusted_width = min(max_length + 3, 50)  # Máximo 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    logging.getLogger(__name__).info(f"  ✓ Formato aplicado a '{worksheet.title}'")


def insertar_imagen_en_excel(
    worksheet,
    imagen_path: Union[str, Path],
    celda: str,
    escala: float = 1.0
) -> None:
    """
    Inserta una imagen PNG en una celda específica de Excel.

    Args:
        worksheet: Objeto worksheet de openpyxl
        imagen_path: Ruta de la imagen PNG
        celda: Celda donde anclar la imagen (ej: 'A1')
        escala: Factor de escala para el tamaño (default: 1.0)

    Ejemplo:
        >>> insertar_imagen_en_excel(ws, "graficos/ventas.png", "B2", escala=0.8)
    """
    imagen_path = Path(imagen_path)

    if not imagen_path.exists():
        raise FileNotFoundError(f"Imagen no encontrada: {imagen_path}")

    img = XLImage(str(imagen_path))

    # Ajustar tamaño
    if escala != 1.0:
        img.width = int(img.width * escala)
        img.height = int(img.height * escala)

    # Insertar en la celda
    worksheet.add_image(img, celda)

    logging.getLogger(__name__).info(f"  ✓ Imagen insertada en {celda}")


def crear_tabla_excel(
    worksheet,
    rango: str,
    nombre_tabla: str,
    estilo: str = 'TableStyleMedium9'
) -> None:
    """
    Convierte un rango de celdas en una tabla formateada de Excel.

    Args:
        worksheet: Objeto worksheet de openpyxl
        rango: Rango de celdas (ej: 'A1:D100')
        nombre_tabla: Nombre único para la tabla
        estilo: Estilo de tabla de Excel (default: TableStyleMedium9 - azul)

    Ejemplo:
        >>> crear_tabla_excel(ws, 'A1:G500', 'TablaVentas')
    """
    tabla = Table(displayName=nombre_tabla, ref=rango)

    # Aplicar estilo
    style = TableStyleInfo(
        name=estilo,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = style

    worksheet.add_table(tabla)

    logging.getLogger(__name__).info(f"  ✓ Tabla '{nombre_tabla}' creada en rango {rango}")


# ═══════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ═══════════════════════════════════════════════════════════════

def formatear_moneda_columna(worksheet, columna_letra: str, fila_inicio: int = 2):
    """
    Aplica formato de moneda ($X,XXX.XX) a una columna completa.

    Args:
        worksheet: Objeto worksheet de openpyxl
        columna_letra: Letra de la columna (ej: 'D')
        fila_inicio: Fila donde inicia el formato (default: 2, después del header)
    """
    max_row = worksheet.max_row

    for row in range(fila_inicio, max_row + 1):
        celda = f"{columna_letra}{row}"
        worksheet[celda].number_format = '$#,##0.00'


def formatear_porcentaje_columna(worksheet, columna_letra: str, fila_inicio: int = 2):
    """
    Aplica formato de porcentaje (XX.XX%) a una columna completa.

    Args:
        worksheet: Objeto worksheet de openpyxl
        columna_letra: Letra de la columna (ej: 'E')
        fila_inicio: Fila donde inicia el formato (default: 2)
    """
    max_row = worksheet.max_row

    for row in range(fila_inicio, max_row + 1):
        celda = f"{columna_letra}{row}"
        worksheet[celda].number_format = '0.00%'


def formatear_numero_columna(worksheet, columna_letra: str, fila_inicio: int = 2):
    """
    Aplica formato numérico con separadores de miles a una columna.

    Args:
        worksheet: Objeto worksheet de openpyxl
        columna_letra: Letra de la columna (ej: 'C')
        fila_inicio: Fila donde inicia el formato (default: 2)
    """
    max_row = worksheet.max_row

    for row in range(fila_inicio, max_row + 1):
        celda = f"{columna_letra}{row}"
        worksheet[celda].number_format = '#,##0'


# ═══════════════════════════════════════════════════════════════
# INICIALIZACIÓN DEL MÓDULO
# ═══════════════════════════════════════════════════════════════

# Configurar matplotlib para mejor apariencia
plt.rcParams['figure.facecolor'] = 'white'
plt.rcParams['axes.facecolor'] = 'white'
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans']

# Logger del módulo
logger = setup_logger(__name__)
