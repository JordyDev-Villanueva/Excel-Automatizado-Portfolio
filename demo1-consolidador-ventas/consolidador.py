"""
Consolidador de Ventas - Demo 1
================================
Script profesional que consolida múltiples archivos Excel de ventas
en un único reporte con análisis avanzados y visualizaciones.

Características:
- Lectura automática de múltiples archivos Excel
- Consolidación y limpieza de datos
- Análisis de ventas por sucursal, producto, vendedor y categoría
- Generación de gráficos profesionales de alta calidad
- Creación de reporte Excel multi-hoja con formato corporativo

Autor: Excel Automatizado
Fecha: Diciembre 2024
"""

import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar encoding para Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding='utf-8')

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE RUTAS
# ═══════════════════════════════════════════════════════════════

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
SHARED_UTILS = BASE_DIR.parent / "shared_utils"

# Agregar shared_utils al path
sys.path.insert(0, str(SHARED_UTILS))

# Importar utilidades compartidas
from excel_helper import (
    setup_logger,
    leer_archivos_excel,
    crear_grafico_barras,
    crear_grafico_circular,
    crear_grafico_linea,
    aplicar_formato_profesional,
    insertar_imagen_en_excel,
    crear_tabla_excel,
    formatear_moneda_columna,
    formatear_numero_columna,
    formatear_porcentaje_columna,
    COLORES
)

# ═══════════════════════════════════════════════════════════════
# LOGGER
# ═══════════════════════════════════════════════════════════════

logger = setup_logger(__name__)

# ═══════════════════════════════════════════════════════════════
# COLUMNAS REQUERIDAS
# ═══════════════════════════════════════════════════════════════

COLUMNAS_REQUERIDAS = [
    'Fecha', 'Producto', 'Categoría', 'Cantidad',
    'Precio_Unitario', 'Vendedor', 'Sucursal'
]

# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE ANÁLISIS
# ═══════════════════════════════════════════════════════════════

def calcular_estadisticas_generales(df: pd.DataFrame) -> dict:
    """
    Calcula estadísticas generales del consolidado.

    Args:
        df: DataFrame consolidado

    Returns:
        dict: Diccionario con métricas clave
    """
    stats = {
        'total_ventas': df['Total_Venta'].sum(),
        'total_transacciones': len(df),
        'ticket_promedio': df['Total_Venta'].mean(),
        'fecha_inicio': df['Fecha'].min(),
        'fecha_fin': df['Fecha'].max(),
        'num_sucursales': df['Sucursal'].nunique(),
        'num_vendedores': df['Vendedor'].nunique(),
        'num_productos': df['Producto'].nunique()
    }

    return stats


def analizar_ventas_por_sucursal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analiza ventas agregadas por sucursal.

    Args:
        df: DataFrame consolidado

    Returns:
        pd.DataFrame: Análisis por sucursal con participación porcentual
    """
    resumen = df.groupby('Sucursal').agg({
        'Total_Venta': 'sum',
        'Fecha': 'count'  # Cuenta transacciones
    }).round(2)

    resumen.columns = ['Total_Ventas', 'Transacciones']

    # Calcular participación porcentual
    resumen['Participacion'] = (resumen['Total_Ventas'] / resumen['Total_Ventas'].sum())

    # Ordenar por ventas descendente
    resumen = resumen.sort_values('Total_Ventas', ascending=False)

    return resumen


def analizar_top_productos(df: pd.DataFrame, top_n: int = 10) -> tuple:
    """
    Identifica los productos más vendidos y más rentables.

    Args:
        df: DataFrame consolidado
        top_n: Cantidad de productos en el top (default: 10)

    Returns:
        tuple: (top_por_cantidad, top_por_monto)
    """
    # Top por cantidad vendida
    top_cantidad = df.groupby('Producto')['Cantidad'].sum().nlargest(top_n)

    # Top por monto total
    top_monto = df.groupby('Producto')['Total_Venta'].sum().nlargest(top_n).round(2)

    return top_cantidad, top_monto


def analizar_vendedores(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analiza el desempeño de cada vendedor.

    Args:
        df: DataFrame consolidado

    Returns:
        pd.DataFrame: Métricas por vendedor
    """
    vendedores = df.groupby('Vendedor').agg({
        'Total_Venta': 'sum',
        'Fecha': 'count'
    }).round(2)

    vendedores.columns = ['Total_Ventas', 'Transacciones']

    # Calcular ticket promedio
    vendedores['Ticket_Promedio'] = (vendedores['Total_Ventas'] / vendedores['Transacciones']).round(2)

    # Ordenar por ventas totales
    vendedores = vendedores.sort_values('Total_Ventas', ascending=False)

    return vendedores


def analizar_categorias(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analiza ventas por categoría de producto.

    Args:
        df: DataFrame consolidado

    Returns:
        pd.DataFrame: Ventas por categoría
    """
    categorias = df.groupby('Categoría').agg({
        'Total_Venta': 'sum',
        'Cantidad': 'sum'
    }).round(2)

    categorias.columns = ['Total_Ventas', 'Unidades_Vendidas']

    # Calcular participación
    categorias['Participacion'] = (categorias['Total_Ventas'] / categorias['Total_Ventas'].sum())

    categorias = categorias.sort_values('Total_Ventas', ascending=False)

    return categorias


def analizar_tendencia_temporal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Analiza la tendencia de ventas en el tiempo.

    Args:
        df: DataFrame consolidado

    Returns:
        pd.DataFrame: Ventas agrupadas por fecha
    """
    # Asegurar que Fecha sea datetime
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    tendencia = df.groupby('Fecha')['Total_Venta'].sum().round(2)

    return tendencia


# ═══════════════════════════════════════════════════════════════
# FUNCIÓN PARA CREAR DASHBOARD
# ═══════════════════════════════════════════════════════════════

def crear_hoja_dashboard(worksheet, stats: dict, graficos_dir: Path):
    """
    Crea la hoja Dashboard con KPIs y gráficos insertados.

    Args:
        worksheet: Hoja de Excel donde crear el dashboard
        stats: Diccionario con estadísticas generales
        graficos_dir: Ruta donde están los gráficos PNG
    """
    # Configurar anchos de columna
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 35
    worksheet.column_dimensions['C'].width = 20

    # ───────────────────────────────────────────────────────────
    # TÍTULO PRINCIPAL
    # ───────────────────────────────────────────────────────────
    worksheet['B2'] = '📊 REPORTE CONSOLIDADO DE VENTAS'
    worksheet['B2'].font = Font(name='Calibri', size=16, bold=True, color=COLORES['azul_oscuro'])
    worksheet['B2'].alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[2].height = 30

    # Fecha del reporte
    worksheet['B3'] = f"Período: {stats['fecha_inicio']} al {stats['fecha_fin']}"
    worksheet['B3'].font = Font(name='Calibri', size=10, color=COLORES['gris'])

    # ───────────────────────────────────────────────────────────
    # KPIs PRINCIPALES
    # ───────────────────────────────────────────────────────────
    fila_kpi = 5

    kpis = [
        ('Total Ventas:', f"${stats['total_ventas']:,.2f}"),
        ('Total Transacciones:', f"{stats['total_transacciones']:,}"),
        ('Ticket Promedio:', f"${stats['ticket_promedio']:,.2f}"),
        ('Sucursales:', f"{stats['num_sucursales']}"),
        ('Vendedores:', f"{stats['num_vendedores']}"),
        ('Productos Únicos:', f"{stats['num_productos']}")
    ]

    for i, (etiqueta, valor) in enumerate(kpis):
        fila = fila_kpi + i

        # Etiqueta
        worksheet[f'B{fila}'] = etiqueta
        worksheet[f'B{fila}'].font = Font(name='Calibri', size=11, bold=True)
        worksheet[f'B{fila}'].alignment = Alignment(horizontal='left')

        # Valor
        worksheet[f'C{fila}'] = valor
        worksheet[f'C{fila}'].font = Font(name='Calibri', size=11, color=COLORES['verde'])
        worksheet[f'C{fila}'].fill = PatternFill(start_color=COLORES['azul_claro'],
                                                  end_color=COLORES['azul_claro'],
                                                  fill_type='solid')
        worksheet[f'C{fila}'].alignment = Alignment(horizontal='center')

        worksheet.row_dimensions[fila].height = 25

    # ───────────────────────────────────────────────────────────
    # GRÁFICOS
    # ───────────────────────────────────────────────────────────

    # Gráfico 1: Ventas por Sucursal (lado izquierdo)
    fila_grafico1 = 13
    worksheet[f'B{fila_grafico1}'] = '📊 Ventas por Sucursal'
    worksheet[f'B{fila_grafico1}'].font = Font(name='Calibri', size=12, bold=True, color=COLORES['azul_oscuro'])

    grafico1_path = graficos_dir / "ventas_sucursal.png"
    if grafico1_path.exists():
        insertar_imagen_en_excel(worksheet, grafico1_path, f'B{fila_grafico1 + 1}', escala=0.55)

    # Gráfico 2: Distribución por Categoría (lado derecho)
    fila_grafico2 = 13
    worksheet[f'I{fila_grafico2}'] = '🥧 Distribución por Categoría'
    worksheet[f'I{fila_grafico2}'].font = Font(name='Calibri', size=12, bold=True, color=COLORES['azul_oscuro'])

    grafico2_path = graficos_dir / "categorias.png"
    if grafico2_path.exists():
        insertar_imagen_en_excel(worksheet, grafico2_path, f'I{fila_grafico2 + 1}', escala=0.55)

    # Gráfico 3: Tendencia Temporal (abajo, centrado)
    fila_grafico3 = 38
    worksheet[f'B{fila_grafico3}'] = '📈 Tendencia de Ventas Diarias'
    worksheet[f'B{fila_grafico3}'].font = Font(name='Calibri', size=12, bold=True, color=COLORES['azul_oscuro'])

    grafico3_path = graficos_dir / "tendencia.png"
    if grafico3_path.exists():
        insertar_imagen_en_excel(worksheet, grafico3_path, f'B{fila_grafico3 + 1}', escala=0.6)

    logger.info("  ✓ Dashboard creado con KPIs y gráficos")


# ═══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def main():
    """
    Función principal que ejecuta el consolidador.
    """
    logger.info("=" * 60)
    logger.info("CONSOLIDADOR DE VENTAS - DEMO 1")
    logger.info("=" * 60)
    logger.info("")

    try:
        # ═══════════════════════════════════════════════════════
        # PASO 1: LEER Y CONSOLIDAR ARCHIVOS
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 1: Leyendo archivos Excel...")

        df_consolidado = leer_archivos_excel(
            INPUT_DIR,
            patron="*.xlsx",
            columnas_requeridas=COLUMNAS_REQUERIDAS
        )

        logger.info(f"✓ {len(df_consolidado)} registros consolidados")
        logger.info("")

        # ═══════════════════════════════════════════════════════
        # PASO 2: CALCULAR COLUMNA TOTAL_VENTA
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 2: Calculando totales...")

        df_consolidado['Total_Venta'] = (
            df_consolidado['Cantidad'] * df_consolidado['Precio_Unitario']
        ).round(2)

        logger.info("✓ Columna 'Total_Venta' calculada")
        logger.info("")

        # ═══════════════════════════════════════════════════════
        # PASO 3: REALIZAR ANÁLISIS
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 3: Realizando análisis...")

        # Estadísticas generales
        stats = calcular_estadisticas_generales(df_consolidado)
        logger.info(f"  → Total ventas: ${stats['total_ventas']:,.2f}")

        # Análisis por sucursal
        ventas_sucursal = analizar_ventas_por_sucursal(df_consolidado)
        logger.info(f"  → Análisis de {len(ventas_sucursal)} sucursales")

        # Top productos
        top_cantidad, top_monto = analizar_top_productos(df_consolidado, top_n=10)
        logger.info(f"  → Top 10 productos identificados")

        # Análisis de vendedores
        vendedores_stats = analizar_vendedores(df_consolidado)
        logger.info(f"  → Desempeño de {len(vendedores_stats)} vendedores analizado")

        # Análisis por categoría
        categorias_stats = analizar_categorias(df_consolidado)
        logger.info(f"  → {len(categorias_stats)} categorías analizadas")

        # Tendencia temporal
        tendencia = analizar_tendencia_temporal(df_consolidado)
        logger.info(f"  → Tendencia de {len(tendencia)} días calculada")

        logger.info("✓ Análisis completado")
        logger.info("")

        # ═══════════════════════════════════════════════════════
        # PASO 4: GENERAR GRÁFICOS
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 4: Generando gráficos profesionales...")

        # Crear carpeta para gráficos temporales
        OUTPUT_DIR.mkdir(exist_ok=True)
        GRAFICOS_DIR = OUTPUT_DIR / "graficos_temp"
        GRAFICOS_DIR.mkdir(exist_ok=True)

        # Gráfico 1: Ventas por Sucursal
        crear_grafico_barras(
            data=ventas_sucursal['Total_Ventas'],
            titulo="Ventas Totales por Sucursal",
            output_path=GRAFICOS_DIR / "ventas_sucursal.png",
            formato_y='miles'
        )

        # Gráfico 2: Distribución por Categoría
        crear_grafico_circular(
            data=categorias_stats['Total_Ventas'],
            titulo="Distribución de Ventas por Categoría",
            output_path=GRAFICOS_DIR / "categorias.png"
        )

        # Gráfico 3: Tendencia Temporal
        crear_grafico_linea(
            data=tendencia,
            titulo="Evolución de Ventas Diarias",
            output_path=GRAFICOS_DIR / "tendencia.png"
        )

        logger.info("✓ Gráficos generados")
        logger.info("")

        # ═══════════════════════════════════════════════════════
        # PASO 5: CREAR ARCHIVO EXCEL CONSOLIDADO
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 5: Creando archivo Excel consolidado...")

        output_file = OUTPUT_DIR / "reporte_consolidado.xlsx"

        # Crear Excel con pandas ExcelWriter
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # HOJA 1: Dashboard (vacía por ahora, la llenaremos después)
            pd.DataFrame().to_excel(writer, sheet_name='Dashboard', index=False)

            # HOJA 2: Datos Consolidados
            df_consolidado.to_excel(writer, sheet_name='Datos_Consolidados', index=False)

            # HOJA 3: Top Productos
            # Crear DataFrame combinado con las dos tablas
            top_productos_df = pd.DataFrame({
                'Producto_Cantidad': top_cantidad.index,
                'Cantidad': top_cantidad.values,
                'Espaciador': [''] * len(top_cantidad),  # Columna vacía
                'Producto_Monto': top_monto.index,
                'Total_Venta': top_monto.values
            })
            top_productos_df.to_excel(writer, sheet_name='Top_Productos', index=False)

            # HOJA 4: Análisis Vendedores
            vendedores_stats.to_excel(writer, sheet_name='Analisis_Vendedores')

            # HOJA 5: Resumen Sucursales
            ventas_sucursal.to_excel(writer, sheet_name='Resumen_Sucursales')

        logger.info("✓ Estructura del Excel creada")

        # ═══════════════════════════════════════════════════════
        # PASO 6: APLICAR FORMATOS PROFESIONALES
        # ═══════════════════════════════════════════════════════
        logger.info("Paso 6: Aplicando formatos profesionales...")

        # Abrir el workbook para formatear
        wb = load_workbook(output_file)

        # ───────────────────────────────────────────────────────
        # FORMATEAR HOJA 1: Dashboard
        # ───────────────────────────────────────────────────────
        ws_dashboard = wb['Dashboard']
        crear_hoja_dashboard(ws_dashboard, stats, GRAFICOS_DIR)

        # ───────────────────────────────────────────────────────
        # FORMATEAR HOJA 2: Datos Consolidados
        # ───────────────────────────────────────────────────────
        ws_datos = wb['Datos_Consolidados']
        aplicar_formato_profesional(ws_datos)

        # Formatear columnas específicas
        formatear_moneda_columna(ws_datos, 'E')  # Precio_Unitario
        formatear_moneda_columna(ws_datos, 'H')  # Total_Venta
        formatear_numero_columna(ws_datos, 'D')  # Cantidad

        # Crear tabla
        max_row_datos = ws_datos.max_row
        crear_tabla_excel(ws_datos, f'A1:H{max_row_datos}', 'TablaDatosConsolidados')

        # ───────────────────────────────────────────────────────
        # FORMATEAR HOJA 3: Top Productos
        # ───────────────────────────────────────────────────────
        ws_top = wb['Top_Productos']

        # Títulos personalizados
        ws_top['A1'] = 'MÁS VENDIDOS (Por Cantidad)'
        ws_top['D1'] = 'MÁS RENTABLES (Por Monto)'

        for cell_ref in ['A1', 'D1']:
            ws_top[cell_ref].font = Font(name='Calibri', size=12, bold=True, color=COLORES['blanco'])
            ws_top[cell_ref].fill = PatternFill(start_color=COLORES['azul_oscuro'],
                                                 end_color=COLORES['azul_oscuro'],
                                                 fill_type='solid')
            ws_top[cell_ref].alignment = Alignment(horizontal='center')

        aplicar_formato_profesional(ws_top, tiene_header=True)
        formatear_numero_columna(ws_top, 'B', fila_inicio=2)  # Cantidad
        formatear_moneda_columna(ws_top, 'E', fila_inicio=2)  # Total_Venta

        # ───────────────────────────────────────────────────────
        # FORMATEAR HOJA 4: Análisis Vendedores
        # ───────────────────────────────────────────────────────
        ws_vendedores = wb['Analisis_Vendedores']
        aplicar_formato_profesional(ws_vendedores)

        formatear_moneda_columna(ws_vendedores, 'B', fila_inicio=2)  # Total_Ventas
        formatear_numero_columna(ws_vendedores, 'C', fila_inicio=2)  # Transacciones
        formatear_moneda_columna(ws_vendedores, 'D', fila_inicio=2)  # Ticket_Promedio

        max_row_vendedores = ws_vendedores.max_row
        crear_tabla_excel(ws_vendedores, f'A1:D{max_row_vendedores}', 'TablaVendedores')

        # ───────────────────────────────────────────────────────
        # FORMATEAR HOJA 5: Resumen Sucursales
        # ───────────────────────────────────────────────────────
        ws_sucursales = wb['Resumen_Sucursales']
        aplicar_formato_profesional(ws_sucursales)

        formatear_moneda_columna(ws_sucursales, 'B', fila_inicio=2)  # Total_Ventas
        formatear_numero_columna(ws_sucursales, 'C', fila_inicio=2)  # Transacciones
        formatear_porcentaje_columna(ws_sucursales, 'D', fila_inicio=2)  # Participación

        # Agregar fila de totales
        ultima_fila = ws_sucursales.max_row + 1
        ws_sucursales[f'A{ultima_fila}'] = 'TOTAL'
        ws_sucursales[f'B{ultima_fila}'] = f'=SUM(B2:B{ultima_fila-1})'
        ws_sucursales[f'C{ultima_fila}'] = f'=SUM(C2:C{ultima_fila-1})'
        ws_sucursales[f'D{ultima_fila}'] = f'=SUM(D2:D{ultima_fila-1})'

        # Formato a la fila de totales
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_sucursales[f'{col}{ultima_fila}']
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.fill = PatternFill(start_color=COLORES['azul_claro'],
                                   end_color=COLORES['azul_claro'],
                                   fill_type='solid')

        formatear_moneda_columna(ws_sucursales, 'B', fila_inicio=ultima_fila)
        formatear_numero_columna(ws_sucursales, 'C', fila_inicio=ultima_fila)
        formatear_porcentaje_columna(ws_sucursales, 'D', fila_inicio=ultima_fila)

        # ───────────────────────────────────────────────────────
        # GUARDAR WORKBOOK
        # ───────────────────────────────────────────────────────
        wb.save(output_file)

        logger.info("✓ Formatos aplicados exitosamente")
        logger.info("")

        # ═══════════════════════════════════════════════════════
        # RESUMEN FINAL
        # ═══════════════════════════════════════════════════════
        logger.info("=" * 60)
        logger.info("✓ ¡REPORTE CONSOLIDADO GENERADO EXITOSAMENTE!")
        logger.info("=" * 60)
        logger.info(f"📁 Archivo: {output_file.name}")
        logger.info(f"📊 Ubicación: {output_file}")
        logger.info("")
        logger.info("RESUMEN:")
        logger.info(f"  • Total Ventas: ${stats['total_ventas']:,.2f}")
        logger.info(f"  • Transacciones: {stats['total_transacciones']:,}")
        logger.info(f"  • Ticket Promedio: ${stats['ticket_promedio']:,.2f}")
        logger.info(f"  • Período: {stats['fecha_inicio']} al {stats['fecha_fin']}")
        logger.info(f"  • Sucursales: {stats['num_sucursales']}")
        logger.info(f"  • Vendedores: {stats['num_vendedores']}")
        logger.info("=" * 60)

    except Exception as e:
        logger.error("=" * 60)
        logger.error(f"❌ ERROR: {str(e)}")
        logger.error("=" * 60)
        raise


# ═══════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
